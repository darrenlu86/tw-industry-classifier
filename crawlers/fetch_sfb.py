#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
fetch_sfb.py — 金管會證券期貨局（SFB／證期局）權威名冊 fetcher（交付版，可獨立執行、可重跑）

本檔為交付版：抓取／解析邏輯與來源 URL 逐字沿用已驗證的產線版本，僅把路徑收斂到
`_paths` 模組、把寫死的抓取日期改成 `--run-date` 參數，不再依賴 原產線的目錄結構。

資料來源
--------
1. 證券商 broker      —— TWSE OpenAPI：https://openapi.twse.com.tw/v1/brokerService/brokerList
2. 期貨商 futures     —— 證期局「期貨業名冊」Excel（外國期貨商／期貨自營商／期貨經紀商）
3. 期貨業四小類 futures_misc —— 同一份「期貨業名冊」Excel（期貨經理／期貨顧問／期貨信託／槓桿交易商）
4. 投信（專營）sitc   —— 證期局「投信公司名冊」Excel
5. 投顧（專營）sica   —— 證期局「投顧公司名冊」Excel
6. 證券金融 sfc       —— 證期局「證金公司名冊」Excel

2-6 的下載連結優先**動態**從證期局名冊頁解析（`resolve_fsc_links()`）：
    https://www.sfb.gov.tw/ch/home.jsp?id=1015&parentpath=0,4
    （「業務主題專區」>「證券期貨特許事業」>「證券期貨業者名冊」）
避免檔名／日期版本更新後連結失效；解析失敗的項目自動落回 `FALLBACK_FSC_URLS`
（2026-07-03 調查紀錄中已驗證的 fsc.gov.tw 固定連結，2026-07-14 重新開啟名冊頁核對
仍為現行版本；其中證金名冊頁面雖標示 115.7.1，檔名仍延用舊檔名「更新 110_7_31證金名錄(1).xlsx」，
此為 FSC 網站自身命名不一致，非本爬蟲誤植）。

輸出
----
- 六份快照 CSV 寫到 `_paths.raw_dir(run_date)`（即 `crawlers/raw/<執行日>/`）：
  sfb_broker.csv／sfb_futures.csv／sfb_futures_misc.csv／sfb_sitc.csv／sfb_sica.csv／sfb_sfc.csv
- 異動報告 `sfb_diff_report.txt` 同目錄。
- 欄位固定 7 欄：segment, inst_code, official_name, org_type, source_url, fetched_at_note, name_note
  （不含 tax_id；統編補全屬下游 enrich_tax_id 職責，不在本檔範圍）

比對基準（baseline）
-------------------
`crawlers/baseline/sfb_*.csv`，由 `_paths.baseline("檔名.csv")` 取得。

已知覆蓋率限制（誠實揭露，非隱藏）
--------------------------------
TWSE OpenAPI brokerList 回傳之 `Name` 為交易所慣用簡稱（如「合庫」＝合作金庫證券
股份有限公司），非公司登記全銜；且其 `Code`（如 1020）與證期局 Excel 名冊的
「證券商代號」（如 0200）是兩套完全不同的編碼系統。因此證券商改用 TWSE 後：
  (a) official_name 欄為簡稱，非全銜——如需全銜／統編，屬下游統編補全（GCIS 商工登記
      比對）之職責，不在本檔產出中處理。
  (b) 與 Excel-based 的舊快照（代碼／名稱皆不同編碼系統）逐碼／逐名比對「新增／消失」
      並無意義；本檔對 broker 只回報列數差異並標註此限制，不假裝兩者可逐筆對齊。
另投顧名冊含專營／兼營兩區塊，本檔僅保留專營；期貨業名冊之「期貨交易輔助人」區塊
刻意排除，不納入任何 segment。

用法
----
  py -3.12 -X utf8 fetch_sfb.py
  py -3.12 -X utf8 fetch_sfb.py --run-date 2026-07-30
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _paths

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# 常數
# ---------------------------------------------------------------------------

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; industry-classifier-fetch-sfb/1.0)"}

SFB_ROSTER_PAGE = "https://www.sfb.gov.tw/ch/home.jsp?id=1015&parentpath=0,4"
TWSE_BROKER_API = "https://openapi.twse.com.tw/v1/brokerService/brokerList"

# 2026-07-03 調查紀錄中已驗證過的固定連結。
# resolve_fsc_links() 動態解析失敗時的 fallback，不是本檔唯一依據。
FALLBACK_FSC_URLS = {
    "futures": "https://www.fsc.gov.tw/userfiles/file/1150701%E6%9C%9F%E8%B2%A8%E6%A5%AD%E5%90%8D%E5%86%8A.xls",
    "sitc": "https://www.fsc.gov.tw/userfiles/file/115_07_01%E6%8A%95%E4%BF%A1%E5%90%8D%E5%86%8A.xlsx",
    "sica": "https://www.fsc.gov.tw/userfiles/file/115_07_01%E6%8A%95%E9%A1%A7%E5%90%8D%E5%86%8A.xlsx",
    "sfc": "https://www.fsc.gov.tw/userfiles/file/%E6%9B%B4%E6%96%B0%20110_7_31%E8%AD%89%E9%87%91%E5%90%8D%E9%8C%84(1).xlsx",
}
# 名冊頁的 row label -> segment key（用於動態解析 EXCEL 連結）
FSC_ROW_LABELS = {
    "期貨業名冊": "futures",
    "投信公司名冊": "sitc",
    "投顧公司名冊": "sica",
    "證金公司名冊": "sfc",
}

SCHEMA = ["segment", "inst_code", "official_name", "org_type", "source_url", "fetched_at_note", "name_note"]


# ---------------------------------------------------------------------------
# 抓取：requests 優先，Playwright 為 fallback（本檔所有來源實測皆 server-rendered
# 靜態內容／官方檔案直連，未觸發 Playwright fallback；仍保留手段以應對日後變化）
# ---------------------------------------------------------------------------


def _get_bytes_via_requests(url: str, timeout: int = 30) -> bytes:
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r.content


def _get_bytes_via_playwright(url: str, timeout_ms: int = 30000) -> bytes:
    """requests 失敗時的 fallback：用 Playwright 的 request context 直接抓 binary/JSON。"""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            ctx = browser.new_context()
            resp = ctx.request.get(url, timeout=timeout_ms, headers=HEADERS)
            if not resp.ok:
                raise RuntimeError(f"Playwright request 非 2xx：status={resp.status}")
            body = resp.body()
        finally:
            browser.close()
    return body


def fetch_bytes(url: str, label: str) -> tuple[bytes, list[str]]:
    """requests 優先，失敗才 fallback 到 Playwright；兩者皆失敗則丟出例外。"""
    attempts: list[str] = []
    try:
        data = _get_bytes_via_requests(url)
        attempts.append(f"[OK] requests：{label}（{len(data)} bytes）")
        return data, attempts
    except Exception as e:  # noqa: BLE001 - 有意寬鬆捕捉，記錄後嘗試下一手段
        attempts.append(f"[FAIL] requests：{label} — {e!r}")
    try:
        data = _get_bytes_via_playwright(url)
        attempts.append(f"[OK] playwright fallback：{label}（{len(data)} bytes）")
        return data, attempts
    except Exception as e:  # noqa: BLE001
        attempts.append(f"[FAIL] playwright fallback：{label} — {e!r}")
    raise RuntimeError(f"fetch_bytes 兩種手段皆失敗（{label}）：" + " ｜ ".join(attempts))


# ---------------------------------------------------------------------------
# 動態解析證期局名冊頁的 Excel 連結（避免檔名/版本更新後 fallback 常數失效）
# ---------------------------------------------------------------------------


def resolve_fsc_links() -> tuple[dict[str, str], list[str]]:
    """回傳 {segment_key: 現行 EXCEL 連結}；解析失敗的項目自動落回 FALLBACK_FSC_URLS。"""
    notes: list[str] = []
    resolved: dict[str, str] = {}
    try:
        html_bytes, attempts = fetch_bytes(SFB_ROSTER_PAGE, "sfb_roster_page")
        notes.extend(attempts)
        html = html_bytes.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "html.parser")
        for tr in soup.find_all("tr"):
            tds = tr.find_all("td")
            if not tds:
                continue
            row_text = " | ".join(td.get_text(strip=True) for td in tds)
            for label, seg_key in FSC_ROW_LABELS.items():
                if label in row_text and seg_key not in resolved:
                    links = tr.find_all("a", href=True)
                    excel_hrefs = [a["href"] for a in links if a.get_text(strip=True).upper() == "EXCEL"]
                    if excel_hrefs:
                        resolved[seg_key] = excel_hrefs[0]
                        notes.append(f"動態解析 {label} -> {seg_key}：{excel_hrefs[0]}")
    except Exception as e:  # noqa: BLE001
        notes.append(f"[FAIL] 動態解析名冊頁失敗，全數落回 fallback 固定連結：{e!r}")

    for seg_key, fallback_url in FALLBACK_FSC_URLS.items():
        if seg_key not in resolved:
            resolved[seg_key] = fallback_url
            notes.append(f"{seg_key} 使用 fallback 固定連結（2026-07-03 調查紀錄）：{fallback_url}")
        elif resolved[seg_key] != fallback_url:
            notes.append(
                f"注意：{seg_key} 動態解析連結與 fallback 常數不同！"
                f"動態={resolved[seg_key]} fallback={fallback_url}（採用動態解析結果，"
                "代表官網已更新檔案，請人工確認是否需要同步更新 FALLBACK_FSC_URLS 常數）"
            )
    return resolved, notes


# ---------------------------------------------------------------------------
# 共用小工具
# ---------------------------------------------------------------------------


def clean_code(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip().replace("\xa0", "").strip()


def split_name(raw) -> tuple[str, str | None]:
    """回傳 (主名稱, 備註) —— 名冊中部分名稱在換行後夾帶沿革說明。"""
    if raw is None:
        return "", None
    s = str(raw).strip()
    if "\n" in s:
        head, rest = s.split("\n", 1)
        return head.strip(), rest.strip()
    return s, None


# ---------------------------------------------------------------------------
# 1. 證券商 broker —— TWSE OpenAPI brokerService/brokerList
# ---------------------------------------------------------------------------


def fetch_broker_twse(fetched_at: str) -> tuple[list[dict], list[str]]:
    notes: list[str] = []
    data, attempts = fetch_bytes(TWSE_BROKER_API, "TWSE brokerList")
    notes.extend(attempts)
    records = json.loads(data.decode("utf-8"))
    notes.append(f"TWSE brokerList 回傳 {len(records)} 筆")

    fetched_at_note = (
        f"{fetched_at} 呼叫 TWSE OpenAPI brokerService/brokerList（{TWSE_BROKER_API}，"
        "即時 JSON API，非人工下載 Excel 後解析）取得現行證券商代號名單。"
        "限制：Name 為交易所慣用簡稱（如「合庫」＝合作金庫證券股份有限公司），"
        "非公司登記全銜；Code 與證期局 Excel 名冊之「證券商代號」為不同編碼系統；"
        "全銜與統編查得屬下游統編補全（GCIS 商工登記比對）之職責，"
        "不在本檔產出範圍。"
    )
    rows = []
    for rec in records:
        code = clean_code(rec.get("Code"))
        name = str(rec.get("Name") or "").strip()
        if not name:
            continue
        rows.append({
            "segment": "broker",
            "inst_code": code,
            "official_name": name,
            "org_type": "證券商（TWSE OpenAPI brokerList，Name 為簡稱非完整法定名稱）",
            "source_url": TWSE_BROKER_API,
            "fetched_at_note": fetched_at_note,
            "name_note": (
                f"地址：{rec.get('Address', '')}；電話：{rec.get('Telephone', '')}；"
                f"核准日期(民國)：{rec.get('EstablishmentDate', '')}"
            ),
        })
    return rows, notes


# ---------------------------------------------------------------------------
# 2/3. 期貨業名冊 -> futures（期貨商） + futures_misc（期貨業四小類）
# ---------------------------------------------------------------------------

FUTURES_SEGMENT_SECTIONS = {"外國期貨商", "期貨自營商", "期貨經紀商"}
FUTURES_MISC_SECTIONS = {"期貨經理事業", "期貨顧問事業", "期貨信託事業", "槓桿交易商"}
FUTURES_EXCLUDED_SECTIONS = {"期貨交易輔助人"}
FUTURES_ALL_SECTIONS = FUTURES_SEGMENT_SECTIONS | FUTURES_MISC_SECTIONS | FUTURES_EXCLUDED_SECTIONS


def _code_prefix_note(code: str) -> str:
    if not code:
        return ""
    p = code[0].upper()
    return {
        "F": "F碼-期貨經紀商本業代號",
        "P": "P碼-期貨自營商代號",
        "S": "S碼-證券商兼營代號",
        "R": "R碼-外國期貨商代號",
        "M": "M碼-期貨經理事業代號",
        "A": "A碼-期貨顧問事業代號",
        "T": "T碼-期貨信託事業代號",
        "L": "L碼-槓桿交易商代號",
        "B": "B碼-期貨交易輔助人代號",
    }.get(p, "")


def parse_futures_xls(xls_bytes: bytes, source_url: str, fetched_at_note: str) -> tuple[list[dict], list[dict]]:
    if xlrd is None:
        raise RuntimeError("xlrd 未安裝，無法解析 .xls 期貨業名冊")
    wb = xlrd.open_workbook(file_contents=xls_bytes)
    sheet_names = wb.sheet_names()
    sh = wb.sheet_by_name("Sheet1") if "Sheet1" in sheet_names else wb.sheet_by_index(0)

    section = None
    fut_rows, misc_rows = [], []
    for r in range(sh.nrows):
        row = sh.row_values(r)
        while row and (row[-1] == "" or row[-1] is None):
            row.pop()
        if not row:
            continue
        c0 = str(row[0]).strip() if row[0] != "" else ""
        c2 = str(row[2]).strip() if len(row) > 2 else ""
        if c0 == "項次" and len(row) > 2:
            for key in FUTURES_ALL_SECTIONS:
                if c2.startswith(key):
                    section = key
                    break
            continue
        if c0.startswith("★") or c0.startswith("臺灣期貨交易所") or c0.startswith("https://") or c0.startswith("中華民國期貨業"):
            continue
        if section is None:
            continue
        code = clean_code(row[1]) if len(row) > 1 else ""
        name_raw = row[2] if len(row) > 2 else ""
        name, note = split_name(str(name_raw).replace("\xa0", "").strip())
        if not name:
            continue
        code = code.replace("\xa0", "").strip()
        cp = _code_prefix_note(code)
        org_type = section + (f"（{cp}）" if cp else "")
        rec = {
            "inst_code": code,
            "official_name": name,
            "org_type": org_type,
            "source_url": source_url,
            "fetched_at_note": fetched_at_note,
            "name_note": note or "",
        }
        if section in FUTURES_SEGMENT_SECTIONS:
            rec["segment"] = "futures"
            fut_rows.append(rec)
        elif section in FUTURES_MISC_SECTIONS:
            rec["segment"] = "futures_misc"
            misc_rows.append(rec)
        elif section in FUTURES_EXCLUDED_SECTIONS:
            continue
    return fut_rows, misc_rows


# ---------------------------------------------------------------------------
# 4. 投信名錄 -> sitc
# ---------------------------------------------------------------------------


def parse_sitc_xlsx(xlsx_bytes: bytes, source_url: str, fetched_at_note: str) -> list[dict]:
    if openpyxl is None:
        raise RuntimeError("openpyxl 未安裝，無法解析投信公司名冊 .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["投信"] if "投信" in wb.sheetnames else wb.worksheets[0]
    out = []
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        while row and row[-1] is None:
            row.pop()
        if not row:
            continue
        c0 = row[0]
        if c0 in ("項次",) or (isinstance(c0, str) and (c0.startswith("★") or c0.startswith("http"))):
            continue
        if not isinstance(c0, (int, float)):
            continue
        code = clean_code(row[1]) if len(row) > 1 else ""
        name_raw = row[2] if len(row) > 2 else ""
        name, note = split_name(name_raw)
        if not name:
            continue
        out.append({
            "segment": "sitc",
            "inst_code": code,
            "official_name": name,
            "org_type": "證券投資信託事業(專營)",
            "source_url": source_url,
            "fetched_at_note": fetched_at_note,
            "name_note": note or "",
        })
    return out


# ---------------------------------------------------------------------------
# 5. 投顧名錄 -> sica（含專營/兼營兩區塊，僅保留專營）
# ---------------------------------------------------------------------------


def parse_sica_xlsx(xlsx_bytes: bytes, source_url: str, fetched_at_note: str) -> list[dict]:
    if openpyxl is None:
        raise RuntimeError("openpyxl 未安裝，無法解析投顧公司名冊 .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["投顧"] if "投顧" in wb.sheetnames else wb.worksheets[0]
    out = []
    section = None
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        while row and row[-1] is None:
            row.pop()
        if not row:
            continue
        c0 = row[0]
        c2 = row[2] if len(row) > 2 else ""
        if c0 == "項次" and isinstance(c2, str):
            section = "兼營" if "兼營投顧事業" in c2 else "專營"
            continue
        if isinstance(c0, str) and (c0.startswith("★") or c0.startswith("http")):
            continue
        if not isinstance(c0, (int, float)):
            continue
        if section is None:
            continue
        code = clean_code(row[1]) if len(row) > 1 else ""
        name_raw = row[2] if len(row) > 2 else ""
        name, note = split_name(name_raw)
        if not name:
            continue
        org_type = "證券投資顧問事業(專營)" if section == "專營" else "兼營投顧事業之公司(投信/銀行/證券/期貨商兼營，非專營投顧)"
        out.append({
            "segment": "sica",
            "inst_code": code,
            "official_name": name,
            "org_type": org_type,
            "source_url": source_url,
            "fetched_at_note": fetched_at_note,
            "name_note": note or "",
        })
    return out


# ---------------------------------------------------------------------------
# 6. 證金名錄 -> sfc
# ---------------------------------------------------------------------------


def parse_sfc_xlsx(xlsx_bytes: bytes, source_url: str, fetched_at_note: str) -> list[dict]:
    if openpyxl is None:
        raise RuntimeError("openpyxl 未安裝，無法解析證金公司名冊 .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["證金"] if "證金" in wb.sheetnames else wb.worksheets[0]
    out = []
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        while row and row[-1] is None:
            row.pop()
        if not row:
            continue
        c0 = row[0]
        if not isinstance(c0, (int, float)):
            continue
        code = clean_code(c0)
        name_raw = row[1] if len(row) > 1 else ""
        name, note = split_name(name_raw)
        if not name:
            continue
        out.append({
            "segment": "sfc",
            "inst_code": code,
            "official_name": name,
            "org_type": "證券金融事業",
            "source_url": source_url,
            "fetched_at_note": fetched_at_note + "（檔名沿用舊檔名「110_7_31」，惟頁面標示更新日 115.7.1，內容為現行唯一持牌證金公司）",
            "name_note": note or "",
        })
    return out


# ---------------------------------------------------------------------------
# CSV I/O
# ---------------------------------------------------------------------------


def write_csv(path: str, rows: list[dict]) -> None:
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SCHEMA)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in SCHEMA})


def read_existing_csv(path: str) -> list[dict] | None:
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Diff：今日快照 vs baseline/ 既有快照
# ---------------------------------------------------------------------------


def diff_roster(name: str, new_rows: list[dict], baseline_path: str, key_caveat: str | None = None) -> dict:
    baseline_rows = read_existing_csv(baseline_path)
    if baseline_rows is None:
        return {
            "name": name,
            "rows": len(new_rows),
            "baseline_rows": None,
            "text": f"## {name}\n無既有快照可比對（baseline 不存在：{baseline_path}）",
        }

    baseline_by_code = {r["inst_code"]: r["official_name"] for r in baseline_rows}
    new_by_code = {r["inst_code"]: r["official_name"] for r in new_rows}

    lines = [
        f"## {name}",
        f"既有快照 {len(baseline_rows)} 列 → 今日快照 {len(new_rows)} 列（差 {len(new_rows) - len(baseline_rows):+d}）",
    ]
    if key_caveat:
        lines.append(f"**注意**：{key_caveat}")
        return {
            "name": name,
            "rows": len(new_rows),
            "baseline_rows": len(baseline_rows),
            "text": "\n".join(lines),
        }

    added = sorted(set(new_by_code) - set(baseline_by_code))
    removed = sorted(set(baseline_by_code) - set(new_by_code))
    renamed = [
        (code, baseline_by_code[code], new_by_code[code])
        for code in (set(new_by_code) & set(baseline_by_code))
        if baseline_by_code[code] != new_by_code[code]
    ]

    lines.append(
        "新增機構：無" if not added
        else f"新增機構 {len(added)} 筆：" + "; ".join(f"{c}={new_by_code[c]}" for c in added)
    )
    lines.append(
        "消失機構：無" if not removed
        else f"消失機構 {len(removed)} 筆：" + "; ".join(f"{c}={baseline_by_code[c]}" for c in removed)
    )
    lines.append(
        "名稱變更：無" if not renamed
        else f"名稱變更 {len(renamed)} 筆：" + "; ".join(f"{c}: {old}→{new}" for c, old, new in renamed)
    )

    return {
        "name": name,
        "rows": len(new_rows),
        "baseline_rows": len(baseline_rows),
        "text": "\n".join(lines),
    }


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="金管會證期局權威名冊 fetcher（交付版）")
    ap.add_argument(
        "--run-date",
        default=date.today().isoformat(),
        help="抓取日（yyyy-mm-dd），決定輸出目錄 raw/<日期>/ 與 fetched_at 註記；預設今天",
    )
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    fetched_at = args.run_date
    raw_dir = _paths.raw_dir(fetched_at)

    all_notes: list[str] = []
    issues: list[str] = []
    # (report_name, rows, output_path, baseline_path, key_caveat)
    results: list[tuple[str, list[dict], str, str, str | None]] = []

    # 0) 動態解析證期局名冊頁的 Excel 連結（futures/sitc/sica/sfc 共用）
    fsc_urls, resolve_notes = resolve_fsc_links()
    all_notes.extend(resolve_notes)

    # 1) 證券商 broker —— TWSE OpenAPI（見檔頭說明）
    try:
        broker_rows, notes = fetch_broker_twse(fetched_at)
        all_notes.extend(notes)
        broker_path = os.path.join(raw_dir, "sfb_broker.csv")
        write_csv(broker_path, broker_rows)
        results.append((
            "sfb_broker（證券商）", broker_rows, broker_path, _paths.baseline("sfb_broker.csv"),
            "改用 TWSE OpenAPI brokerList 取代原 Excel 名冊：inst_code 編碼系統與 official_name "
            "命名慣例（簡稱 vs 全銜）皆與既有 Excel-based 快照不同，逐碼/逐名比對「新增/消失」無意義。"
            "本次快照可作為往後 TWSE-based 快照互相比較的新基期。",
        ))
    except Exception as e:  # noqa: BLE001
        issues.append(
            f"證券商(broker) TWSE OpenAPI 抓取失敗：{e!r}（manual fallback：沿用既有快照 "
            "baseline/sfb_broker.csv；需人工至 "
            f"{TWSE_BROKER_API} 或 www.sfb.gov.tw 名冊頁核對）"
        )

    # 2/3) 期貨業名冊 -> futures + futures_misc
    try:
        futures_url = fsc_urls["futures"]
        xls_bytes, notes = fetch_bytes(futures_url, "期貨業名冊.xls")
        all_notes.extend(notes)
        fetched_note = (
            f"{fetched_at} 從 {SFB_ROSTER_PAGE}（『業務主題專區』>『證券期貨特許事業』>"
            f"『證券期貨業者名冊』）取得『期貨業名冊』連結（{futures_url}），"
            "requests 直接下載 .xls 後以 xlrd 解析。"
        )
        fut_rows, misc_rows = parse_futures_xls(xls_bytes, futures_url, fetched_note)
        all_notes.append(f"期貨業名冊解析：futures={len(fut_rows)} 筆，futures_misc={len(misc_rows)} 筆")

        futures_path = os.path.join(raw_dir, "sfb_futures.csv")
        misc_path = os.path.join(raw_dir, "sfb_futures_misc.csv")
        write_csv(futures_path, fut_rows)
        write_csv(misc_path, misc_rows)
        results.append((
            "sfb_futures（期貨商）", fut_rows, futures_path, _paths.baseline("sfb_futures.csv"), None,
        ))
        results.append((
            "sfb_futures_misc（期貨業四小類原始名冊列，未去重兼營前）",
            misc_rows, misc_path, _paths.baseline("sfb_futures_misc.csv"), None,
        ))
    except Exception as e:  # noqa: BLE001
        issues.append(
            f"期貨業名冊抓取/解析失敗：{e!r}（manual fallback：沿用既有快照 "
            "baseline/sfb_futures.csv 與 baseline/sfb_futures_misc.csv；"
            f"需人工至 {SFB_ROSTER_PAGE} 核對『期貨業名冊』連結）"
        )

    # 4) 投信名錄 -> sitc
    try:
        sitc_url = fsc_urls["sitc"]
        xlsx_bytes, notes = fetch_bytes(sitc_url, "投信公司名冊.xlsx")
        all_notes.extend(notes)
        fetched_note = (
            f"{fetched_at} 從 {SFB_ROSTER_PAGE} 取得『投信公司名冊』連結（{sitc_url}），"
            "requests 直接下載 .xlsx 後以 openpyxl 解析。"
        )
        sitc_rows = parse_sitc_xlsx(xlsx_bytes, sitc_url, fetched_note)
        all_notes.append(f"投信公司名冊解析：sitc={len(sitc_rows)} 筆")
        sitc_path = os.path.join(raw_dir, "sfb_sitc.csv")
        write_csv(sitc_path, sitc_rows)
        results.append((
            "sfb_sitc（投信）", sitc_rows, sitc_path, _paths.baseline("sfb_sitc.csv"), None,
        ))
    except Exception as e:  # noqa: BLE001
        issues.append(
            f"投信公司名冊抓取/解析失敗：{e!r}（manual fallback：沿用既有快照 "
            f"baseline/sfb_sitc.csv；需人工至 {SFB_ROSTER_PAGE} 核對）"
        )

    # 5) 投顧名錄 -> sica
    try:
        sica_url = fsc_urls["sica"]
        xlsx_bytes, notes = fetch_bytes(sica_url, "投顧公司名冊.xlsx")
        all_notes.extend(notes)
        fetched_note = (
            f"{fetched_at} 從 {SFB_ROSTER_PAGE} 取得『投顧公司名冊』連結（{sica_url}），"
            "requests 直接下載 .xlsx 後以 openpyxl 解析（僅保留專營投顧區塊）。"
        )
        sica_rows = parse_sica_xlsx(xlsx_bytes, sica_url, fetched_note)
        all_notes.append(f"投顧公司名冊解析：sica(專營)={len(sica_rows)} 筆")
        sica_path = os.path.join(raw_dir, "sfb_sica.csv")
        write_csv(sica_path, sica_rows)
        results.append((
            "sfb_sica（投顧-專營）", sica_rows, sica_path, _paths.baseline("sfb_sica.csv"), None,
        ))
    except Exception as e:  # noqa: BLE001
        issues.append(
            f"投顧公司名冊抓取/解析失敗：{e!r}（manual fallback：沿用既有快照 "
            f"baseline/sfb_sica.csv；需人工至 {SFB_ROSTER_PAGE} 核對）"
        )

    # 6) 證金名錄 -> sfc
    try:
        sfc_url = fsc_urls["sfc"]
        xlsx_bytes, notes = fetch_bytes(sfc_url, "證金公司名冊.xlsx")
        all_notes.extend(notes)
        fetched_note = (
            f"{fetched_at} 從 {SFB_ROSTER_PAGE} 取得『證金公司名冊』連結（{sfc_url}），"
            "requests 直接下載 .xlsx 後以 openpyxl 解析。"
        )
        sfc_rows = parse_sfc_xlsx(xlsx_bytes, sfc_url, fetched_note)
        all_notes.append(f"證金公司名冊解析：sfc={len(sfc_rows)} 筆")
        sfc_path = os.path.join(raw_dir, "sfb_sfc.csv")
        write_csv(sfc_path, sfc_rows)
        results.append((
            "sfb_sfc（證券金融）", sfc_rows, sfc_path, _paths.baseline("sfb_sfc.csv"), None,
        ))
    except Exception as e:  # noqa: BLE001
        issues.append(
            f"證金公司名冊抓取/解析失敗：{e!r}（manual fallback：沿用既有快照 "
            f"baseline/sfb_sfc.csv；需人工至 {SFB_ROSTER_PAGE} 核對）"
        )

    # ---- Diff 報告 ----
    report_lines = [f"# fetch_sfb 抓取與 diff 報告 — {fetched_at}", ""]
    report_lines.append("## 抓取過程紀錄")
    report_lines.extend(f"- {n}" for n in all_notes)
    report_lines.append("")

    for name, rows, _path, baseline_path, caveat in results:
        d = diff_roster(name, rows, baseline_path, key_caveat=caveat)
        report_lines.append(d["text"])
        report_lines.append("")

    if issues:
        report_lines.append("## 失敗紀錄（manual fallback）")
        report_lines.extend(f"- {i}" for i in issues)
        report_lines.append("")

    print("\n".join(report_lines))

    if results:
        os.makedirs(raw_dir, exist_ok=True)
        report_path = os.path.join(raw_dir, "sfb_diff_report.txt")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(report_lines))
        print(f"[written] {report_path}")

    return 1 if issues else 0


if __name__ == "__main__":
    sys.exit(main())
