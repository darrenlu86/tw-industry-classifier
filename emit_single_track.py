# -*- coding: utf-8 -*-
r"""產出「單軌版」產業分類對照表（xlsx）— 每戶只有一組 產業大類／產業子類

單軌合併規則（2026-08-11 使用者定版；合併邏輯在 core/engine.py 的 query()，
本檔只做排版輸出）：

    身分軌命中（金控與銀行／證券期貨／保險／租賃／電支支付／政府機關／教育與法人）
        → 產業大類＝身分大類、產業子類＝身分子類
    一般企業（官方名冊查無執照）
        → 產業大類＝行業大類（稅務行業標準分類第 9 次修訂 A–S 十九大類）
          產業子類＝行業中類（2 碼＋名稱）；稅籍查無時依序再試 L2-5 上市櫃名冊、
          L3-D GCIS 登記狀態（→「已解散（歷史戶）」），皆未命中才標「未登記（稅籍查無）」
    L0 終結（v5：個人戶／境外法人（無台灣登記）／無法分類）與執行業務者（非營業人）
        → 沿用身分軌值，產業子類多為空白

欄位樣貌同 產業分類對照表_20260803.xlsx（六欄），列序與輸入一致，可並排比對。

用法：
    py -3.12 emit_single_track.py
        讀 output/產業分類對照表_20260803.xlsx（取「統一編號」與「原始資料產業別」欄），
        寫 output/產業分類對照表_<今日>_單軌版.xlsx
    py -3.12 emit_single_track.py --in <輸入.xlsx> --out <輸出.xlsx>

輸出兩個分頁：
    產業分類對照表  主表六欄，列序與輸入一致（可並排比對）
    周邊單位        L1-2 白名單全列（內建＋本地例外檔追加），供複核

產物含客戶清單 → 一律寫到 output/（已被 .gitignore 排除，不會進公開 remote）。
"""
import argparse
import os
import sys
import time
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "core"))

import engine                                        # noqa: E402
import rules as R                                    # noqa: E402
import exceptions as X                               # noqa: E402
from classify import build_provider                  # noqa: E402
from emit_eight_column import read_input             # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

OUTPUT = os.path.join(HERE, "output")
DEFAULT_IN = os.path.join(OUTPUT, "產業分類對照表_20260803.xlsx")

HEADERS = ["統一編號", "官方正式名稱", "產業大類", "產業子類",
           "分類依據詞", "原始資料產業別"]

# 名冊與稅籍皆查無時的依據詞：引擎的兜底語（「兜底歸一般企業」）與名稱關鍵字語
# 描述的是身分層，單軌表已不呈現身分層 → 改用直述查無的通用語（L1 人工裁決除外）。
TAX_MISSING_BASIS = "名冊查無執照且稅籍查無"


def six_columns(provider, raw_tax_id, raw_industry=""):
    """一筆輸入 → 六欄 list。"""
    rec = engine.query(raw_tax_id, provider)
    basis = rec["分類依據詞"].rstrip()               # 引擎對超長裁決理由截斷，去尾端空白
    if (rec["大分類"] == "一般企業"
            and rec["產業大類"] == R.TAX_MISSING_SECTION
            and not rec["分類依據層"].startswith("L1")):
        basis = TAX_MISSING_BASIS
    return [rec["統一編號"], rec["官方正式名稱"],
            rec["產業大類"], rec["產業子類"], basis, raw_industry]


def write_xlsx(path, main_rows, peri_rows):
    from openpyxl import Workbook                    # noqa: PLC0415
    from openpyxl.styles import Font                 # noqa: PLC0415
    wb = Workbook()
    widths = [12, 44, 26, 26, 46, 14]

    ws = wb.active
    ws.title = "產業分類對照表"
    ws.append(HEADERS)
    for row in main_rows:
        ws.append(row)
    ws2 = wb.create_sheet("周邊單位")
    ws2.append(HEADERS[:5] + ["名單來源"])
    for row in peri_rows:
        ws2.append(row)
    for sheet in (ws, ws2):
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = w
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions      # 表頭下拉篩選（沿 0803 版樣貌）
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="產出單軌版產業分類對照表（產業大類＋產業子類）")
    ap.add_argument("--in", dest="src", default=DEFAULT_IN,
                    help="輸入 xlsx（需含「統一編號」欄；預設 %s）" % os.path.basename(DEFAULT_IN))
    ap.add_argument("--out", dest="dst", default="",
                    help="輸出 xlsx（預設 output/產業分類對照表_<今日>_單軌版.xlsx）")
    ap.add_argument("--mode", choices=("auto", "local", "api"), default="local")
    ap.add_argument("--offline", action="store_true", help="不連外（GCIS 名稱解析會跳過）")
    args = ap.parse_args()

    if not os.path.exists(args.src):
        sys.exit("找不到輸入檔：%s" % args.src)
    items = read_input(args.src)
    mode, provider = build_provider(args.mode, args.offline)
    print("模式 %s ｜ 輸入 %d 筆（%s）" % (mode, len(items), os.path.basename(args.src)))
    if mode == "local" and getattr(provider, "wants_preload", False):
        print("預載稅籍（掃一次全檔，約 20–40 秒）…")
        provider.preload([engine.normalize_tax_id(t) for t, _ in items])

    main_rows = [six_columns(provider, tid, raw) for tid, raw in items]

    peri_rows = []
    for tid in sorted(X.PERIPHERAL):
        row = six_columns(provider, tid)
        peri_rows.append(row[:5] + ["內建白名單" if tid in X.PERIPHERAL_BUILTIN
                                    else "本地例外檔"])

    if hasattr(provider, "save_gcis_cache"):
        provider.save_gcis_cache()

    os.makedirs(OUTPUT, exist_ok=True)
    dst = args.dst or os.path.join(
        OUTPUT, "產業分類對照表_%s_單軌版.xlsx" % time.strftime("%Y%m%d"))
    write_xlsx(dst, main_rows, peri_rows)

    print()
    for k, v in Counter(r[2] for r in main_rows).most_common():
        print("  %-24s %4d" % (k, v))
    print()
    n_builtin = len(set(X.PERIPHERAL) & set(X.PERIPHERAL_BUILTIN))
    print("周邊單位分頁：%d 列（內建 %d＋本地 %d）"
          % (len(peri_rows), n_builtin, len(peri_rows) - n_builtin))
    print("寫出 %s（主表 %d 列）" % (dst, len(main_rows)))


if __name__ == "__main__":
    main()
