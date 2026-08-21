# -*- coding: utf-8 -*-
r"""產出「八欄版」產業分類對照表（xlsx）— 兩軌並列

八欄兩軌（2026-08-03 提案架構；2026-08-11 使用者核定調整後由本檔落地，
取代原留在工作階段暫存區、未進版控的一次性產表腳本）：

    身分軌  身分大類／身分子類 —— 引擎判定（core/rules.py v4）。
            呈現規則：「一般企業」子類留空、依據詞改為通用查無語
            （身分層的細分由行業軌回答；標籤沿用「一般企業」，
            「未列管」改名案經使用者 2026-08-11 裁示不採）。
    行業軌  行業大類（A–S 十九大類）／行業中類（2 碼＋名稱）——
            稅籍主行業代號直查 rules.SECTION_BY_MAJOR2，禁止人工裁決。

用法：
    py -3.12 emit_eight_column.py
        讀 output/產業分類對照表_20260803.xlsx（取「統一編號」與「原始資料產業別」欄），
        寫 output/產業分類對照表_<今日>_八欄版.xlsx
    py -3.12 emit_eight_column.py --in <輸入.xlsx> --out <輸出.xlsx>

輸出兩個分頁：
    產業分類對照表  主表八欄，列序與輸入一致（可並排比對）
    周邊單位        L1-2 白名單全列（內建＋本地例外檔追加），供複核

產物含客戶清單 → 一律寫到 output/（已被 .gitignore 排除，不會進公開 remote）。
"""
import argparse
import os
import sys
import time
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "core"))

import engine                                        # noqa: E402
import exceptions as X                               # noqa: E402
from classify import build_provider                  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

OUTPUT = os.path.join(HERE, "output")
DEFAULT_IN = os.path.join(OUTPUT, "產業分類對照表_20260803.xlsx")

HEADERS = ["統一編號", "官方正式名稱", "身分大類", "身分子類",
           "行業大類", "行業中類", "分類依據詞", "原始資料產業別"]

# 一般企業呈現：身分層查無執照者不帶自訂子類，產業資訊由行業軌回答。
# 標籤維持「一般企業」（使用者 2026-08-11 裁示；2026-08-03 提案的「未列管」改名不採）。
UNTRACKED_LABEL = "一般企業"
UNTRACKED_BASIS = "名冊查無：金管會／機關／學校／非營利名冊皆未命中，且非稅籍特許碼"


def eight_columns(provider, raw_tax_id, raw_industry=""):
    """一筆輸入 → 八欄 list。"""
    rec = engine.query(raw_tax_id, provider)
    tid = rec["統一編號"]                             # 輸入原值（L1-1 不改寫輸出鍵）
    if rec["分類依據層"].startswith("L0"):            # L0 終結者沒有 8 碼統編，行業軌不適用
        section, major2 = "", ""
    else:                                             # 行業軌查詢比照引擎走歸戶後的統編
        section, major2 = engine.industry_track(provider, engine.lookup_tax_id(tid))
    # 依據詞去尾端空白：引擎對超長裁決理由做 head[:40] 截斷，切點可能落在空白後
    group, sub, basis = rec["大分類"], rec["子分類"], rec["分類依據詞"].rstrip()
    if group == "一般企業":
        group, sub = UNTRACKED_LABEL, ""
        # 人工裁決（L1）落一般企業者保留裁決理由；其餘（稅籍碼／兜底）用通用查無語
        if not rec["分類依據層"].startswith("L1"):
            basis = UNTRACKED_BASIS
    return [tid, rec["官方正式名稱"], group, sub, section, major2, basis, raw_industry]


def read_input(path):
    """讀輸入 xlsx 第一個分頁 → [(統編, 原始資料產業別)]，維持列序。"""
    from openpyxl import load_workbook               # noqa: PLC0415
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    head = [("" if v is None else str(v).strip()) for v in next(rows)]
    if "統一編號" not in head:
        sys.exit("輸入檔缺「統一編號」欄：%s" % path)
    i_tid = head.index("統一編號")
    i_raw = head.index("原始資料產業別") if "原始資料產業別" in head else None
    out = []
    for row in rows:
        tid = "" if row[i_tid] is None else str(row[i_tid]).strip()
        if not tid:
            continue
        raw = ""
        if i_raw is not None and len(row) > i_raw and row[i_raw] is not None:
            raw = str(row[i_raw]).strip()
        out.append((tid, raw))
    wb.close()
    return out


def write_xlsx(path, main_rows, peri_rows):
    from openpyxl import Workbook                    # noqa: PLC0415
    from openpyxl.styles import Font                 # noqa: PLC0415
    wb = Workbook()
    widths = [12, 44, 12, 14, 30, 26, 46, 14]

    ws = wb.active
    ws.title = "產業分類對照表"
    ws.append(HEADERS)
    for row in main_rows:
        ws.append(row)
    ws2 = wb.create_sheet("周邊單位")
    ws2.append(HEADERS[:7] + ["名單來源"])
    for row in peri_rows:
        ws2.append(row)
    for sheet in (ws, ws2):
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = w
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions      # 表頭下拉篩選（沿 0803 版樣貌）
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="產出八欄版產業分類對照表（身分軌＋行業軌）")
    ap.add_argument("--in", dest="src", default=DEFAULT_IN,
                    help="輸入 xlsx（需含「統一編號」欄；預設 %s）" % os.path.basename(DEFAULT_IN))
    ap.add_argument("--out", dest="dst", default="",
                    help="輸出 xlsx（預設 output/產業分類對照表_<今日>_八欄版.xlsx）")
    ap.add_argument("--mode", choices=("auto", "local", "api"), default="local")
    ap.add_argument("--offline", action="store_true", help="不連外（GCIS 名稱解析會跳過）")
    args = ap.parse_args()

    if not os.path.exists(args.src):
        sys.exit("找不到輸入檔：%s" % args.src)
    items = read_input(args.src)
    mode, provider = build_provider(args.mode, args.offline)
    print("模式 %s ｜ 輸入 %d 筆（%s）" % (mode, len(items), os.path.basename(args.src)))
    if mode == "local" and getattr(provider, "wants_preload", False):
        print("預載稅籍（掃一次全檔，約 20–40 秒）…")
        provider.preload([engine.normalize_tax_id(t) for t, _ in items])

    main_rows = [eight_columns(provider, tid, raw) for tid, raw in items]

    peri_rows = []
    for tid in sorted(X.PERIPHERAL):
        row = eight_columns(provider, tid)
        peri_rows.append(row[:7] + ["內建白名單" if tid in X.PERIPHERAL_BUILTIN
                                    else "本地例外檔"])

    if hasattr(provider, "save_gcis_cache"):
        provider.save_gcis_cache()

    os.makedirs(OUTPUT, exist_ok=True)
    dst = args.dst or os.path.join(
        OUTPUT, "產業分類對照表_%s_八欄版.xlsx" % time.strftime("%Y%m%d"))
    write_xlsx(dst, main_rows, peri_rows)

    print()
    for k, v in Counter(r[2] for r in main_rows).most_common():
        print("  %-8s %4d" % (k, v))
    print()
    n_builtin = len(set(X.PERIPHERAL) & set(X.PERIPHERAL_BUILTIN))
    print("周邊單位分頁：%d 列（內建 %d＋本地 %d）"
          % (len(peri_rows), n_builtin, len(peri_rows) - n_builtin))
    print("寫出 %s（主表 %d 列）" % (dst, len(main_rows)))


if __name__ == "__main__":
    main()
